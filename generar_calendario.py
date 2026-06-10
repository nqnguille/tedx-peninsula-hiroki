from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colores brand ──────────────────────────────────────────────
BLACK    = "FF0D0D0D"
RED      = "FFEB0028"
WHITE    = "FFFFFFFF"
GRAY_BG  = "FF1A1A1A"
GRAY_MID = "FF2A2A2A"
GRAY_ROW = "FF141414"
RED_SOFT = "FF3D0008"

PILAR_COLORS = {
    "🧠 Idea":         ("FF1A1A2E", "FFAAB4FF"),
    "🌿 Territorio":   ("FF0D1F14", "FF7FD4A0"),
    "🎤 Convocatoria": ("FF3D0008", "FFFF7088"),
    "💡 Inspiración":  ("FF1F1A00", "FFFFD166"),
    "🎬 Evento":       ("FF001A2E", "FF7EC8E3"),
    "🔧 Behind":       ("FF1A0D1F", "FFCF9FFF"),
}

def pilar_fill(pilar_key):
    bg, _ = PILAR_COLORS.get(pilar_key, ("FF1A1A1A", "FFCCCCCC"))
    return PatternFill("solid", fgColor=bg)

def pilar_font(pilar_key, bold=False):
    _, fg = PILAR_COLORS.get(pilar_key, ("FF1A1A1A", "FFCCCCCC"))
    return Font(name="Calibri", color=fg, bold=bold, size=11)

def header_style(ws, row, col, value, bg=BLACK, fg=WHITE, size=12, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Calibri", color=fg, bold=bold, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def data_cell(ws, row, col, value, bg=GRAY_ROW, fg=WHITE, bold=False, wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Calibri", color=fg, bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c

def thin_border():
    s = Side(style="thin", color="FF333333")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_borders(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()


# ══════════════════════════════════════════════════════════════
# HOJA 1 — PILARES
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📌 Pilares"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = "EB0028"

# Título
ws1.merge_cells("A1:E1")
t = ws1["A1"]
t.value = "TEDxPeninsulaHiroki — Pilares de Contenido"
t.fill = PatternFill("solid", fgColor="FFEB0028")
t.font = Font(name="Calibri", color=WHITE, bold=True, size=16)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 40

# Subtítulo
ws1.merge_cells("A2:E2")
s = ws1["A2"]
s.value = "Cultivando Conciencia · Neuquén 2026"
s.fill = PatternFill("solid", fgColor=BLACK)
s.font = Font(name="Calibri", color="FF888888", italic=True, size=11)
s.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 22

ws1.row_dimensions[3].height = 8

# Headers
headers = ["#", "Pilar", "Qué comunica", "Ejemplos de post", "% del mix"]
cols_w  = [5, 22, 40, 55, 12]
for i, (h, w) in enumerate(zip(headers, cols_w), 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
    header_style(ws1, 4, i, h, bg="FF1A1A1A", fg="FFEB0028", size=11)
ws1.row_dimensions[4].height = 28

pilares = [
    ("1", "🧠 Idea",
     "El tema central del evento: conciencia, atención, IA, presencia.",
     "• La IA puede imitar el pensamiento — ¿qué no puede?\n• 8 segundos de atención promedio en 2026\n• ¿Qué significa estar presente hoy?",
     "25%"),
    ("2", "🌿 Territorio",
     "La reserva Hiroki, la confluencia, la Patagonia. El lugar como argumento.",
     "• Foto aérea de la confluencia\n• Historia de la familia Hiroki\n• Fauna de la reserva (Martín Pescador, Garza mora)\n• El sendero de los dos ríos",
     "20%"),
    ("3", "🎤 Convocatoria",
     "Las 4 convocatorias abiertas: audiencia, speaker, voluntario, partner.",
     "• Carousel Audiencia\n• Carousel Speaker\n• Carousel Voluntario\n• Carousel Partner\n• Recordatorios de cierre",
     "20%"),
    ("4", "💡 Inspiración",
     "Preguntas, citas, ideas que resuenan con el tema. Invitan a pensar.",
     "• Pregunta de la semana (todos los miércoles)\n• Citas de pensadores, científicos, filósofos\n• Recomendaciones de libros/charlas\n• Reflexiones breves",
     "20%"),
    ("5", "🎬 Evento",
     "Qué es TEDx, cómo funciona el formato, qué puede esperar el asistente.",
     "• Qué es TEDx en 5 slides\n• Cómo se seleccionan speakers\n• El formato de las charlas (8-18 min)\n• Las charlas quedan en YouTube para siempre",
     "10%"),
    ("6", "🔧 Behind the scenes",
     "El equipo, el proceso, el making of. Humaniza el evento.",
     "• Primera reunión de curaduría\n• El equipo en acción\n• Preparación del venue\n• Decisiones de diseño del programa",
     "5%"),
]

for r, (num, pilar, que, ejemplos, pct) in enumerate(pilares, 5):
    bg_key = pilar
    ws1.row_dimensions[r].height = 72
    data_cell(ws1, r, 1, num,    bg=GRAY_MID, fg="FF888888", align="center")
    c = ws1.cell(row=r, column=2, value=pilar)
    c.fill = pilar_fill(bg_key)
    c.font = pilar_font(bg_key, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_cell(ws1, r, 3, que,     bg=GRAY_ROW)
    data_cell(ws1, r, 4, ejemplos, bg=GRAY_ROW)
    data_cell(ws1, r, 5, pct,     bg=GRAY_ROW, align="center", fg="FFEB0028", bold=True)

apply_borders(ws1, 4, 10, 1, 5)


# ══════════════════════════════════════════════════════════════
# HOJA 2 — CALENDARIO
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📅 Calendario")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = "1A1A1A"

ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "Calendario de Contenidos — Julio · Agosto · Septiembre 2026"
t2.fill = PatternFill("solid", fgColor="FFEB0028")
t2.font = Font(name="Calibri", color=WHITE, bold=True, size=15)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 38

col_widths = [12, 10, 10, 38, 38, 38, 18]
col_labels = ["Mes", "Semana", "Fecha aprox.", "Post 1 — Lunes", "Post 2 — Miércoles", "Post 3 — Viernes", "Estado"]
for i, (w, lbl) in enumerate(zip(col_widths, col_labels), 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
    header_style(ws2, 2, i, lbl, bg="FF1A1A1A", fg="FFEB0028", size=10)
ws2.row_dimensions[2].height = 26

calendario = [
    # JULIO
    ("JULIO", "Semana 1", "7–11 jul",
     "🎬 Quiénes somos\nCarousel de presentación de la cuenta. Qué es TEDxPeninsulaHiroki, dónde sucede, en qué año.",
     "🌿 La reserva desde el aire\nFoto aérea de la confluencia. Copy: 'Dentro de una ciudad de 400.000 personas, existe un bosque que los dos ríos rodean antes de unirse.'",
     "🧠 Cultivando Conciencia\n¿Qué significa? Slide tipográfico oscuro con la pregunta central del evento.",
     "⬜ Sin hacer"),
    ("JULIO", "Semana 2", "14–18 jul",
     "🌿 La confluencia\nDonde los colores del Limay y el Neuquén se mezclan. El concepto geográfico como metáfora.",
     "🎤 Carousel Audiencia\n'Sos de las personas que escucha una charla y no puede parar de pensar.'",
     "💡 Pregunta del miércoles\n¿Cuándo fue la última vez que una idea te cambió algo?",
     "⬜ Sin hacer"),
    ("JULIO", "Semana 3", "21–25 jul",
     "🎬 Qué es TEDx\n5 slides explicando el formato: charlas de 8-18 min, una sola idea, sin PowerPoints genéricos, van a YouTube.",
     "🌿 Historia Hiroki\nLa familia de agricultores japoneses en la punta de la confluencia. El paisaje que era de una familia, es ahora de todos.",
     "🎤 Carousel Voluntario\n'Te importa más cómo se hacen las cosas que qué se hace.'",
     "⬜ Sin hacer"),
    ("JULIO", "Semana 4", "28 jul–1 ago",
     "💡 Cita inspiracional\nSobre conciencia, atención o presencia. Autor relevante.",
     "🌿 Fauna — Martín Pescador\nFoto + dato de la especie. 'Una señal de que el ecosistema funciona.'",
     "🧠 La pregunta que no es tecnológica\n'La IA puede imitar el pensamiento. ¿Qué queda del juicio propio?'",
     "⬜ Sin hacer"),

    # AGOSTO
    ("AGOSTO", "Semana 1", "4–8 ago",
     "🎤 Carousel Speaker\n'Tenés una idea que se te repite. Una que le contás a todos.'",
     "💡 Pregunta del miércoles\n¿Qué idea te persigue? La que no podés dejar de pensar.",
     "🌿 Foto reserva — sendero\nSendero de otoño entre los sauces. Silencio patagónico.",
     "⬜ Sin hacer"),
    ("AGOSTO", "Semana 2", "11–15 ago",
     "🧠 Dato de atención\n8 segundos de atención promedio. Comparación con 12 segundos en el año 2000.",
     "🎤 Carousel Partner\n'Tu organización cree que las ideas importan más que los logos.'",
     "🔧 El equipo\nFoto o presentación del equipo organizador. Quiénes arman esto y por qué.",
     "⬜ Sin hacer"),
    ("AGOSTO", "Semana 3", "18–22 ago",
     "💡 Cita sobre ideas\nAlguien que habló de cómo las ideas transforman comunidades.",
     "🌿 Fauna — Garza mora\nFoto al amanecer sobre el río. Dato breve.",
     "🎤 Recordatorio Audiencia\nEl cupo es 100. La selección es intencional. ¿Estás anotado/a?",
     "⬜ Sin hacer"),
    ("AGOSTO", "Semana 4", "25–29 ago",
     "🧠 Estar presente — entonces y ahora\n¿Qué significa 'estar presente' en 2026 vs hace 10 años?",
     "🔧 Making of — curaduría\nPrimera reunión del equipo de curaduría de speakers. Sin spoilers.",
     "💡 Pregunta del miércoles\n¿Qué escucharías si pudieras elegir una charla solo para vos?",
     "⬜ Sin hacer"),

    # SEPTIEMBRE
    ("SEPTIEMBRE", "Semana 1", "1–5 sep",
     "🌿 La línea donde se mezclan\nFoto de la confluencia exacta. 'En días claros se ve la línea donde los colores del Limay y el Neuquén empiezan a mezclarse.'",
     "🧠 La frase del manifiesto\n'No es una pregunta tecnológica. Es la más antigua que existe.'",
     "🎤 Último llamado Speaker\nLas postulaciones cierran pronto. Una sola oración es suficiente para empezar.",
     "⬜ Sin hacer"),
    ("SEPTIEMBRE", "Semana 2", "8–12 sep",
     "💡 3 referencias que inspiraron el evento\nLibros, charlas o conceptos. Muestra el marco intelectual del evento.",
     "🔧 Behind the scenes — programa\nEl equipo diseñando el programa del evento. Qué bloques hay, cómo se piensa la experiencia.",
     "🌿 Video corto — reserva\n15 segundos del video de la reserva. Paisaje, agua, árboles. Sin texto, sin logo.",
     "⬜ Sin hacer"),
    ("SEPTIEMBRE", "Semana 3", "15–19 sep",
     "🧠 Carousel — 5 ideas en juego\n5 preguntas que van a estar sobre la mesa el día del evento. Genera anticipación.",
     "💡 Pregunta de la semana\nLa más provocadora del calendario. Para máxima participación.",
     "🎬 El evento en una oración\n'100 personas. Un atardecer. Neuquén.' Imagen del auditorio vacío.",
     "⬜ Sin hacer"),
    ("SEPTIEMBRE", "Semana 4", "22–26 sep",
     "🌿 Atardecer en la reserva\nLa foto más cinematográfica. Cielo, agua, silencio.",
     "🔧 El equipo en el venue\nFotos del equipo preparando el lugar. Última semana antes del evento.",
     "💡 Cita de cierre\nAlgo que deje ganas de estar ahí. El build up final.",
     "⬜ Sin hacer"),
]

MES_COLORS = {
    "JULIO":      ("FF0D1520", "FF7EC8E3"),
    "AGOSTO":     ("FF0D1F14", "FF7FD4A0"),
    "SEPTIEMBRE": ("FF1F0D14", "FFFF7088"),
}

mes_actual = None
for r_off, row_data in enumerate(calendario, 3):
    r = r_off
    ws2.row_dimensions[r].height = 80
    mes, semana, fecha, p1, p2, p3, estado = row_data

    mes_bg, mes_fg = MES_COLORS.get(mes, (BLACK, WHITE))

    if mes != mes_actual:
        mes_actual = mes
        c = ws2.cell(row=r, column=1, value=mes)
        c.fill = PatternFill("solid", fgColor=mes_bg[2:])
        c.font = Font(name="Calibri", color=mes_fg[2:], bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        c = ws2.cell(row=r, column=1, value="")
        c.fill = PatternFill("solid", fgColor=mes_bg[2:])

    data_cell(ws2, r, 2, semana, bg=GRAY_MID, fg="FFAAAAAA", align="center")
    data_cell(ws2, r, 3, fecha,  bg=GRAY_MID, fg="FF666666", align="center")

    # Detectar pilar del post para colorear
    for col, post in [(4, p1), (5, p2), (6, p3)]:
        emoji = post[:2] if post else ""
        pilar_key = next((k for k in PILAR_COLORS if k[:2] == emoji), None)
        c2 = ws2.cell(row=r, column=col, value=post)
        if pilar_key:
            bg_c, fg_c = PILAR_COLORS[pilar_key]
            c2.fill = PatternFill("solid", fgColor=bg_c[2:])
            c2.font = Font(name="Calibri", color=fg_c[2:], size=9)
        else:
            c2.fill = PatternFill("solid", fgColor=GRAY_ROW[2:])
            c2.font = Font(name="Calibri", color=WHITE[2:], size=9)
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Estado
    estado_bg = "FF0D3B0D" if "✅" in estado else "FF1A1A1A"
    data_cell(ws2, r, 7, estado, bg=estado_bg, fg="FF888888", align="center")

apply_borders(ws2, 2, 2 + len(calendario), 1, 7)


# ══════════════════════════════════════════════════════════════
# HOJA 3 — POSTS RECURRENTES
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🔁 Recurrentes")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = "EB0028"

ws3.merge_cells("A1:D1")
t3 = ws3["A1"]
t3.value = "Posts Recurrentes — Templates semanales"
t3.fill = PatternFill("solid", fgColor="FFEB0028")
t3.font = Font(name="Calibri", color=WHITE, bold=True, size=14)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

for i, (w, lbl) in enumerate(zip([28, 15, 45, 45], ["Template", "Frecuencia", "Copy de ejemplo", "Visual sugerida"]), 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
    header_style(ws3, 2, i, lbl, bg="FF1A1A1A", fg="FFEB0028", size=10)
ws3.row_dimensions[2].height = 24

recurrentes = [
    ("💡 Pregunta de la semana", "Todos los miércoles",
     "¿Cuántas veces por día abrís el teléfono sin saber por qué?\n¿Qué harías diferente si supieras que nadie te está mirando?\n¿Cuándo fue la última vez que te aburriste de verdad?\n¿Qué idea cambiaste de opinión este año?\n¿Qué querés que tu hijo piense sobre esta época?",
     "Tipografía blanca grande sobre fondo negro. Sin imagen. La pregunta es el visual."),
    ("🌿 Fauna/flora de la semana", "Todos los viernes",
     "Martín Pescador · Garza mora · Bandurria · Flamenco · Pato cuchara · Sauce llorón · Totora · Pejerrey",
     "Foto de la especie en la reserva. Dato breve en caption: nombre, dónde se ve, qué significa su presencia."),
    ("🧠 Dato de conciencia", "Todos los lunes",
     "En 2004, el tiempo de atención promedio era 2.5 minutos. Hoy: 47 segundos.\nEl 70% de las personas revisa el teléfono dentro de los 4 minutos de despertar.\nLa mente divaga el 47% del tiempo, según Harvard.",
     "Número grande sobre fondo oscuro + fuente citada abajo. Sin adornos."),
    ("📣 Anuncio de nuevo miembro", "Cuando corresponda",
     "✅ Una persona nueva se sumó a la lista de espera.\n✅ Nuevo voluntario en el equipo.\n✅ Partner confirmado.\n✅ Speaker aceptado — 'Tiene una idea sobre [tema].'",
     "Story: fondo negro, nombre o ciudad (si autorizan), ícono del rol en rojo TEDx."),
]

for r_off, (template, freq, copy, visual) in enumerate(recurrentes, 3):
    ws3.row_dimensions[r_off].height = 90
    data_cell(ws3, r_off, 1, template, bg=GRAY_MID, fg="FFEB0028", bold=True)
    data_cell(ws3, r_off, 2, freq, bg=GRAY_ROW, fg="FF888888", align="center")
    data_cell(ws3, r_off, 3, copy, bg=GRAY_ROW)
    data_cell(ws3, r_off, 4, visual, bg=GRAY_ROW, fg="FFAAAAAA")

apply_borders(ws3, 2, 6, 1, 4)


# ══════════════════════════════════════════════════════════════
# HOJA 4 — STORIES CONVOCATORIA
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📲 Stories")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = "333333"

ws4.merge_cells("A1:C1")
t4 = ws4["A1"]
t4.value = "Stories — Anuncios de convocatoria"
t4.fill = PatternFill("solid", fgColor="FF1A1A1A")
t4.font = Font(name="Calibri", color="FFEB0028", bold=True, size=14)
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 36

for i, (w, lbl) in enumerate(zip([22, 42, 42], ["Tipo de anuncio", "Copy sugerido", "Visual"]), 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
    header_style(ws4, 2, i, lbl, bg="FF1A1A1A", fg="FFEB0028", size=10)
ws4.row_dimensions[2].height = 24

stories = [
    ("✅ Nueva persona en lista",
     "Una persona más se anotó para estar en la sala.\n[Ciudad o profesión si autorizan]\n\nEl cupo es 100. Cada lugar cuenta.",
     "Fondo negro. Número acumulado de inscriptos en rojo. Ciudad o perfil anónimo debajo."),
    ("✅ Nuevo voluntario",
     "[Nombre] se suma al equipo de TEDxPeninsulaHiroki.\n[Rol: producción / comunicación / logística]\n\nEl evento se hace entre todos.",
     "Foto del voluntario (si autoriza) o silueta. Nombre y rol en tipografía roja."),
    ("✅ Partner confirmado",
     "[Organización] cree que las ideas importan.\nBienvenidos al equipo.\n\n#TEDxPeninsulaHiroki",
     "Logo del partner sobre fondo negro. Tagline de la organización abajo."),
    ("✅ Speaker aceptado",
     "[Nombre] va a hablar en TEDxPeninsulaHiroki.\nSu idea: [una oración concreta].\n\nPreparate para escucharla.",
     "Foto del speaker (si autoriza). Nombre grande. Su idea en tipografía más pequeña debajo."),
    ("⏳ Recordatorio cierre convocatoria",
     "Quedan [X] días para anotarte.\nEl formulario cierra el [fecha].\n\nLink en bio.",
     "Cuenta regresiva animada (puede hacerse en Canva). Fondo negro, número en rojo."),
]

for r_off, (tipo, copy, visual) in enumerate(stories, 3):
    ws4.row_dimensions[r_off].height = 85
    data_cell(ws4, r_off, 1, tipo, bg=GRAY_MID, fg="FF7FD4A0", bold=True)
    data_cell(ws4, r_off, 2, copy, bg=GRAY_ROW)
    data_cell(ws4, r_off, 3, visual, bg=GRAY_ROW, fg="FFAAAAAA")

apply_borders(ws4, 2, 7, 1, 3)

# ── Guardar ────────────────────────────────────────────────────
output = "/home/guillenqn/proyectos/tedx-peninsula-hiroki/TEDxPeninsulaHiroki_Calendario_Contenidos.xlsx"
wb.save(output)
print(f"✅ Guardado: {output}")
